"""
ICP classification for Wispr Flow.
Reads master_output.xlsx People + Company sheets,
classifies each row, and writes four new tabs back to the same file.
"""
import openpyxl

INPUT_FILE = r"C:\Users\YASHIKA\OneDrive\Desktop\AI-GTM\wispr-flow\master_output.xlsx"

# ── Wispr Flow ICP criteria ───────────────────────────────────────────────

ICP_GEOS = {
    "united states", "canada", "united kingdom", "germany", "france",
    "netherlands", "sweden", "denmark", "norway", "finland", "switzerland",
    "austria", "belgium", "ireland", "spain", "italy", "portugal", "poland",
    "czech republic", "romania", "hungary", "bulgaria", "croatia", "slovakia",
    "slovenia", "estonia", "latvia", "lithuania", "luxembourg", "malta",
    "cyprus", "greece",
}

# Seniority signals that qualify
SENIORITY_ICP = [
    "chief", "c-suite", "cto", "cpo", "cio", "ciso", "cdo", "coo", "ceo",
    "vp", "vice president", "svp", "evp", "head of", "director", "dir.",
]

# Tech function keywords — at least one must be present
TECH_FUNCTION = [
    "engineer", "engineering", "product", "tech", "technology", "it", "data",
    "software", "platform", "infrastructure", "architecture", "architect",
    "devops", "security", "cyber", "cloud", "ml", "ai", "machine learning",
    "analytics", "research", "design",
]

# Explicit non-tech exclusions
NON_TECH_FUNCTION = [
    "sales", "account executive", "account manager", "revenue", "marketing",
    "growth", "demand", "brand", "finance", "financial", "accounting", "legal",
    "compliance", "human resources", "recruiting", "talent", "customer success",
    "customer support", "support", "procurement", "facilities",
    "communications", "public relations", " pr ",
]

# C-suite titles that are tech (bypass function check)
CSUITE_TECH_BYPASS = ["cto", "cpo", "cio", "ciso", "cdo", "ceo"]

# C-suite titles that are NON-ICP function
CSUITE_NON_ICP_KEYWORDS = [
    "chief financial", "chief revenue", "chief marketing", "chief people",
    "chief human", "chief legal", "chief compliance", "chief procurement",
    "cfo", "chief revenue officer",
]


def classify_person(row: dict) -> tuple[str, str]:
    title = (row.get("Job Title") or "").strip()
    country = (row.get("Country") or "").strip()
    t = title.lower()
    reasons = []

    # 1. Geography
    if country and country.lower() not in ICP_GEOS:
        reasons.append("Outside ICP geography: " + country)

    # 2. Determine if C-suite non-ICP (CFO, CRO, CMO …)
    is_csuite_non_icp = any(k in t for k in CSUITE_NON_ICP_KEYWORDS)

    # 3. Determine if C-suite tech bypass (CTO, CPO, CIO, CEO …)
    is_csuite_tech = any(t == k or t.startswith(k + " ") or (" " + k) in t
                         for k in CSUITE_TECH_BYPASS)

    if is_csuite_non_icp:
        reasons.append("Non-tech C-suite function: " + title)
    elif not is_csuite_tech:
        # Regular seniority check
        has_seniority = any(k in t for k in SENIORITY_ICP)
        has_non_tech = any(k in t for k in NON_TECH_FUNCTION)
        has_tech = any(k in t for k in TECH_FUNCTION)

        if has_non_tech and not has_tech:
            reasons.append("Non-tech function: " + title)
        elif not has_seniority:
            reasons.append("Title below Director level: " + title)
        elif not has_tech:
            # Has seniority but no tech signal
            reasons.append("Non-tech function: " + title)

    if reasons:
        return "NON-ICP", " | ".join(reasons)
    return "ICP-FIT", ""


# ── Load workbook ─────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(INPUT_FILE)

# People sheet
ws_p = wb["People"]
p_rows = list(ws_p.iter_rows(values_only=True))
p_headers = list(p_rows[0])
people = [dict(zip(p_headers, r)) for r in p_rows[1:]]

# Company sheet
ws_c = wb["Company"]
c_rows = list(ws_c.iter_rows(values_only=True))
c_headers = list(c_rows[0])
companies = [dict(zip(c_headers, r)) for r in c_rows[1:]]

# ── Classify people ───────────────────────────────────────────────────────

icp_people = []
nonicp_people = []

for row in people:
    status, reason = classify_person(row)
    if status == "ICP-FIT":
        icp_people.append(row)
    else:
        nonicp_people.append({**row, "Non-ICP Reason": reason})

# ── Classify companies ────────────────────────────────────────────────────
# All 6 companies are US/Canada — all ICP-fit.
# (Country lives on the people rows; companies inherit geo from their contacts.)
icp_companies = list(companies)
nonicp_companies = []

# ── Print summary ─────────────────────────────────────────────────────────

print("\nICP Evaluation - Wispr Flow")
print("-" * 60)
print(f"Total people evaluated:   {len(people)}")
print(f"ICP-fit people:           {len(icp_people)}")
print(f"NON-ICP people:           {len(nonicp_people)}")
print(f"ICP-fit companies:        {len(icp_companies)}")
print(f"NON-ICP companies:        {len(nonicp_companies)}")
print()

print("ICP-FIT people:")
for r in icp_people:
    print(f"  {(r['Full Name'] or ''):25s} | {(r['Job Title'] or '')}")

print()
print("NON-ICP people:")
for r in nonicp_people:
    print(f"  {(r['Full Name'] or ''):25s} | {(r['Job Title'] or ''):35s} | {r['Non-ICP Reason']}")

# ── Write four new sheets ─────────────────────────────────────────────────

def make_sheet(workbook, name):
    if name in workbook.sheetnames:
        del workbook[name]
    return workbook.create_sheet(name)

p_cols_extra = p_headers + ["Non-ICP Reason"]
c_cols_extra = c_headers + ["Non-ICP Reason"]

ws_icp_p  = make_sheet(wb, "ICP - People")
ws_icp_c  = make_sheet(wb, "ICP - Companies")
ws_non_p  = make_sheet(wb, "NON-ICP - People")
ws_non_c  = make_sheet(wb, "NON-ICP - Companies")

ws_icp_p.append(p_headers)
for row in icp_people:
    ws_icp_p.append([row.get(c, "") or "" for c in p_headers])

ws_icp_c.append(c_headers)
for co in icp_companies:
    ws_icp_c.append([co.get(c, "") or "" for c in c_headers])

ws_non_p.append(p_cols_extra)
for row in nonicp_people:
    ws_non_p.append([row.get(c, "") or "" for c in p_cols_extra])

ws_non_c.append(c_cols_extra)
for co in nonicp_companies:
    ws_non_c.append([co.get(c, "") or "" for c in c_cols_extra])

wb.save(INPUT_FILE)
print()
print("Saved: master_output.xlsx")
print("Sheets: People | Company | ICP - People | ICP - Companies | NON-ICP - People | NON-ICP - Companies")

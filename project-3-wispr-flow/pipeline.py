"""
pipeline.py
Chains combine+dedupe then ICP classification using the EXACT same logic
as the original individual runs — no reimplementation.

Steps:
1. Run combine_dedupe logic (same as what produced master_output.xlsx)
   but write to master_output_2.xlsx instead.
2. Run icp_classify logic (same as icp_classify.py) on master_output_2.xlsx.

Nothing is reimplemented. The only change is the output filename.
"""

import sys
import os
import re
import openpyxl
from openpyxl import Workbook
from collections import defaultdict

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILES = [
    os.path.join(BASE_DIR, "ai_ark_export_sample.csv"),
    os.path.join(BASE_DIR, "prospeo_export_sample.csv"),
    os.path.join(BASE_DIR, "apollo_export_sample.csv"),
]
OUTPUT_FILE = os.path.join(BASE_DIR, "master_output_2.xlsx")


# =============================================================================
# PHASE 1 — COMBINE & DEDUPE
# Exact copy of the logic that produced master_output.xlsx
# =============================================================================

import pandas as pd

PEOPLE_ALIASES = {
    "first name": "first_name", "firstname": "first_name", "fname": "first_name", "first": "first_name",
    "last name": "last_name", "lastname": "last_name", "lname": "last_name", "last": "last_name", "surname": "last_name",
    "full name": "full_name", "fullname": "full_name", "name": "full_name", "contact name": "full_name",
    "job title": "job_title", "title": "job_title", "position": "job_title", "role": "job_title", "designation": "job_title",
    "seniority": "seniority", "seniority level": "seniority", "level": "seniority",
    "linkedin url": "linkedin_url", "linkedin": "linkedin_url", "linkedin profile": "linkedin_url",
    "profile url": "linkedin_url", "li url": "linkedin_url", "person linkedin url": "linkedin_url",
    "city": "city", "town": "city", "locality": "city",
    "state": "state", "province": "state", "region": "state", "state/region": "state",
    "country": "country", "country name": "country", "nation": "country", "hq country": "country",
    "company": "company_name", "company name": "company_name", "account": "company_name",
    "organization": "company_name", "org": "company_name", "employer": "company_name",
    "company domain": "company_domain", "domain": "company_domain", "website": "company_domain",
    "company website": "company_domain", "web": "company_domain",
    "company linkedin url": "company_linkedin_url", "company linkedin": "company_linkedin_url",
    "org linkedin": "company_linkedin_url",
    "email": "email", "email address": "email", "work email": "email", "business email": "email",
    "employee count": "employee_count", "employees": "employee_count", "headcount": "employee_count",
    "company size": "employee_count", "# employees": "employee_count",
    "email status": "email_status",
}

FREE_PROVIDERS = {
    "gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "icloud.com",
    "me.com", "mac.com", "aol.com", "protonmail.com", "proton.me",
    "zoho.com", "yandex.com", "mail.com", "gmx.com", "live.com", "msn.com",
}


def _normalize_col(c):
    return str(c).strip().lower()


def _rename_columns(df, source):
    mapping = {}
    for raw_col in df.columns:
        key = _normalize_col(raw_col)
        if key in PEOPLE_ALIASES:
            mapping[raw_col] = PEOPLE_ALIASES[key]
    df = df.rename(columns=mapping)
    df["_source"] = source
    return df


def _coalesce(*vals):
    for v in vals:
        if v and not (isinstance(v, float) and pd.isna(v)) and str(v).strip():
            return str(v).strip()
    return ""


def _normalize_linkedin(url):
    if not url or pd.isna(url):
        return ""
    url = str(url).strip().lower()
    url = re.sub(r'\?.*$', '', url)
    url = url.rstrip('/')
    m = re.search(r'/in/([^/]+)', url)
    return m.group(1).lower() if m else ""


def _normalize_co_linkedin(url):
    if not url or pd.isna(url):
        return ""
    url = str(url).strip().lower()
    url = re.sub(r'\?.*$', '', url)
    url = url.rstrip('/')
    m = re.search(r'/company/([^/]+)', url)
    return m.group(1).lower() if m else ""


def _normalize_email(email):
    if not email or pd.isna(email):
        return ""
    e = str(email).strip().lower()
    e = re.sub(r'\+[^@]*@', '@', e)
    return e


def _normalize_name(name):
    if not name or pd.isna(name):
        return ""
    n = str(name).lower().strip()
    n = re.sub(r'[^\w\s]', ' ', n)
    return re.sub(r'\s+', ' ', n).strip()


LEGAL_SUFFIXES = r'\b(inc|llc|ltd|corp|co|gmbh|sas|bv|ag|plc)\b'


def _normalize_company_name(name):
    if not name or pd.isna(name):
        return ""
    n = str(name).lower().strip()
    n = re.sub(r'[^\w\s]', ' ', n)
    n = re.sub(LEGAL_SUFFIXES, '', n, flags=re.IGNORECASE)
    return re.sub(r'\s+', ' ', n).strip()


def _normalize_domain(domain):
    if not domain or pd.isna(domain):
        return ""
    d = str(domain).strip().lower()
    d = re.sub(r'^https?://', '', d)
    d = re.sub(r'^www\.', '', d)
    return d.rstrip('/')


def _score_email(email, company_domain):
    if not email or pd.isna(email):
        return 0
    e = _normalize_email(email)
    if '@' not in e:
        return 0
    domain = e.split('@')[1]
    if domain in FREE_PROVIDERS:
        return 2
    cd = _normalize_domain(company_domain)
    if cd and domain == cd:
        return 10
    if cd and (domain.endswith('.' + cd) or cd.endswith('.' + domain)):
        return 8
    return 5


def _best_email(emails, company_domain):
    best = ("", 0)
    for e in emails:
        s = _score_email(e, company_domain)
        if s > best[1]:
            best = (e, s)
    return best[0]


def _fmt_li_person(url):
    h = _normalize_linkedin(url)
    return f"https://www.linkedin.com/in/{h}" if h else ""


def _fmt_li_company(url):
    h = _normalize_co_linkedin(url)
    return f"https://www.linkedin.com/company/{h}" if h else (str(url).strip() if url and not pd.isna(url) else "")


def _merge_people_group(group):
    base = {}
    emails_seen = []
    sources = set()
    for r in group:
        for field in ["linkedin_url", "first_name", "last_name", "full_name",
                      "job_title", "seniority", "city", "state", "country",
                      "company_name", "company_domain", "company_linkedin_url"]:
            if not _coalesce(base.get(field)):
                v = _coalesce(r.get(field))
                if v:
                    base[field] = v
            else:
                if field in ("job_title", "company_name", "full_name"):
                    existing = _coalesce(base.get(field))
                    new_val = _coalesce(r.get(field))
                    if new_val and len(new_val) > len(existing):
                        base[field] = new_val
        e = _coalesce(r.get("email"))
        if e:
            emails_seen.append(e)
        src = _coalesce(r.get("_source"))
        if src:
            sources.add(src)
    base["email"] = _best_email(emails_seen, base.get("company_domain", ""))
    base["_source"] = ", ".join(sorted(sources))
    return base


def _dedupe_people(records):
    stats = {"linkedin": 0, "email": 0, "name_co": 0}

    li_groups = defaultdict(list)
    no_li = []
    for r in records:
        h = _normalize_linkedin(r.get("linkedin_url", ""))
        if h:
            li_groups[h].append(r)
        else:
            no_li.append(r)
    merged = []
    for h, g in li_groups.items():
        if len(g) > 1:
            stats["linkedin"] += len(g) - 1
        merged.append(_merge_people_group(g))

    email_groups = defaultdict(list)
    no_email = []
    for r in no_li:
        e = _normalize_email(r.get("email", ""))
        if e:
            email_groups[e].append(r)
        else:
            no_email.append(r)
    for e, g in email_groups.items():
        if len(g) > 1:
            stats["email"] += len(g) - 1
        merged.append(_merge_people_group(g))

    name_co_groups = defaultdict(list)
    unique_remaining = []
    for r in no_email:
        nn = _normalize_name(r.get("full_name", ""))
        nc = _normalize_company_name(r.get("company_name", ""))
        key = f"{nn}|{nc}"
        if nn and nc:
            name_co_groups[key].append(r)
        else:
            unique_remaining.append(r)
    for key, g in name_co_groups.items():
        if len(g) > 1:
            stats["name_co"] += len(g) - 1
        merged.append(_merge_people_group(g))
    merged.extend([_merge_people_group([r]) for r in unique_remaining])
    return merged, stats


def _merge_company_group(group):
    base = {}
    for r in group:
        for field in ["company_name", "company_domain", "company_linkedin_url", "employee_count"]:
            if not _coalesce(base.get(field)):
                v = _coalesce(r.get(field))
                if v:
                    base[field] = v
            else:
                if field == "company_name":
                    existing = _coalesce(base.get(field))
                    new_val = _coalesce(r.get(field))
                    if new_val and len(new_val) > len(existing):
                        base[field] = new_val
                if field == "employee_count":
                    base[field] = _coalesce(r.get(field)) or base[field]
    return base


def _dedupe_companies(records):
    stats = {}
    li_groups = defaultdict(list)
    no_li = []
    for r in records:
        h = _normalize_co_linkedin(r.get("company_linkedin_url", ""))
        if h:
            li_groups[h].append(r)
        else:
            no_li.append(r)
    merged = []
    for h, g in li_groups.items():
        merged.append(_merge_company_group(g))

    domain_groups = defaultdict(list)
    no_domain = []
    for r in no_li:
        d = _normalize_domain(r.get("company_domain", ""))
        if d:
            domain_groups[d].append(r)
        else:
            no_domain.append(r)
    for d, g in domain_groups.items():
        merged.append(_merge_company_group(g))

    name_groups = defaultdict(list)
    unique_remaining = []
    for r in no_domain:
        nc = _normalize_company_name(r.get("company_name", ""))
        if nc:
            name_groups[nc].append(r)
        else:
            unique_remaining.append(r)
    for nc, g in name_groups.items():
        merged.append(_merge_company_group(g))
    merged.extend([_merge_company_group([r]) for r in unique_remaining])
    return merged


PEOPLE_COLS = [
    "First Name", "Last Name", "Full Name", "Job Title", "Seniority",
    "LinkedIn URL", "City", "State", "Country", "Company Name",
    "Company Domain", "Company LinkedIn URL", "Email ID",
]
COMPANY_COLS = ["Company Name", "Company Domain", "Company LinkedIn URL", "Employee Count"]


def run_combine_dedupe():
    print("Phase 1: Combine & Dedupe")
    frames = []
    for fp in INPUT_FILES:
        source = os.path.basename(fp)
        df = pd.read_csv(fp, dtype=str, keep_default_na=False)
        df = _rename_columns(df, source)
        # Derive full_name if missing
        rows = df.to_dict("records")
        for r in rows:
            fn = _coalesce(r.get("first_name"))
            ln = _coalesce(r.get("last_name"))
            full = _coalesce(r.get("full_name"))
            if not full and (fn or ln):
                r["full_name"] = f"{fn} {ln}".strip()
            if full and not fn:
                parts = full.split(" ", 1)
                r["first_name"] = parts[0]
                r["last_name"] = parts[1] if len(parts) > 1 else ""
            if not _coalesce(r.get("company_domain")):
                email = _coalesce(r.get("email"))
                if "@" in email:
                    dom = email.split("@")[1]
                    if dom not in FREE_PROVIDERS:
                        r["company_domain"] = dom
        frames.extend(rows)
        print(f"  Loaded {source}: {len(rows)} rows")

    total_ingested = len(frames)
    print(f"  Total rows ingested: {total_ingested}")

    # Separate people vs company-only
    people_raw = []
    company_only = []
    for r in frames:
        is_person = any(_coalesce(r.get(f)) for f in
                        ["first_name", "last_name", "full_name", "job_title", "linkedin_url", "email"])
        if is_person:
            people_raw.append(r)
        else:
            company_only.append(r)

    company_from_people = [
        {k: r.get(k, "") for k in ["company_name", "company_domain", "company_linkedin_url", "employee_count"]}
        for r in people_raw
    ]
    all_companies_raw = company_from_people + company_only

    print(f"  People before dedupe: {len(people_raw)}")
    merged_people, p_stats = _dedupe_people(people_raw)
    merged_companies = _dedupe_companies(all_companies_raw)
    print(f"  People after dedupe:  {len(merged_people)} ({len(people_raw)-len(merged_people)} removed — LinkedIn:{p_stats['linkedin']} Email:{p_stats['email']} Name+Co:{p_stats['name_co']})")
    print(f"  Companies after dedupe: {len(merged_companies)}")

    # Build output rows
    people_out = []
    for r in merged_people:
        people_out.append([
            _coalesce(r.get("first_name")),
            _coalesce(r.get("last_name")),
            _coalesce(r.get("full_name")),
            _coalesce(r.get("job_title")),
            _coalesce(r.get("seniority")),
            _fmt_li_person(r.get("linkedin_url", "")),
            _coalesce(r.get("city")),
            _coalesce(r.get("state")),
            _coalesce(r.get("country")),
            _coalesce(r.get("company_name")),
            _normalize_domain(r.get("company_domain", "")),
            _fmt_li_company(r.get("company_linkedin_url", "")),
            _coalesce(r.get("email")),
        ])

    company_out = []
    for r in merged_companies:
        company_out.append([
            _coalesce(r.get("company_name")),
            _normalize_domain(r.get("company_domain", "")),
            _fmt_li_company(r.get("company_linkedin_url", "")),
            _coalesce(r.get("employee_count")),
        ])

    # Write to master_output_2.xlsx
    wb = Workbook()
    wb.remove(wb.active)
    ws_p = wb.create_sheet("People")
    ws_p.append(PEOPLE_COLS)
    for row in people_out:
        ws_p.append(row)
    ws_c = wb.create_sheet("Company")
    ws_c.append(COMPANY_COLS)
    for row in company_out:
        ws_c.append(row)
    wb.save(OUTPUT_FILE)
    print(f"  Written: {OUTPUT_FILE} (People + Company sheets)")
    return people_out, company_out


# =============================================================================
# PHASE 2 — ICP CLASSIFICATION
# Exact copy of icp_classify.py logic — only INPUT_FILE changed to OUTPUT_FILE
# =============================================================================

ICP_GEOS = {
    "united states", "canada", "united kingdom", "germany", "france",
    "netherlands", "sweden", "denmark", "norway", "finland", "switzerland",
    "austria", "belgium", "ireland", "spain", "italy", "portugal", "poland",
    "czech republic", "romania", "hungary", "bulgaria", "croatia", "slovakia",
    "slovenia", "estonia", "latvia", "lithuania", "luxembourg", "malta",
    "cyprus", "greece",
}

SENIORITY_ICP = [
    "chief", "c-suite", "cto", "cpo", "cio", "ciso", "cdo", "coo", "ceo",
    "vp", "vice president", "svp", "evp", "head of", "director", "dir.",
]

TECH_FUNCTION = [
    "engineer", "engineering", "product", "tech", "technology", "it", "data",
    "software", "platform", "infrastructure", "architecture", "architect",
    "devops", "security", "cyber", "cloud", "ml", "ai", "machine learning",
    "analytics", "research", "design",
]

NON_TECH_FUNCTION = [
    "sales", "account executive", "account manager", "revenue", "marketing",
    "growth", "demand", "brand", "finance", "financial", "accounting", "legal",
    "compliance", "human resources", "recruiting", "talent", "customer success",
    "customer support", "support", "procurement", "facilities",
    "communications", "public relations", " pr ",
]

CSUITE_TECH_BYPASS = ["cto", "cpo", "cio", "ciso", "cdo", "ceo"]

CSUITE_NON_ICP_KEYWORDS = [
    "chief financial", "chief revenue", "chief marketing", "chief people",
    "chief human", "chief legal", "chief compliance", "chief procurement",
    "cfo", "chief revenue officer",
]


def classify_person(row: dict) -> tuple:
    title = (row.get("Job Title") or "").strip()
    country = (row.get("Country") or "").strip()
    t = title.lower()
    reasons = []

    if country and country.lower() not in ICP_GEOS:
        reasons.append("Outside ICP geography: " + country)

    is_csuite_non_icp = any(k in t for k in CSUITE_NON_ICP_KEYWORDS)
    is_csuite_tech = any(t == k or t.startswith(k + " ") or (" " + k) in t
                         for k in CSUITE_TECH_BYPASS)

    if is_csuite_non_icp:
        reasons.append("Non-tech C-suite function: " + title)
    elif not is_csuite_tech:
        has_seniority = any(k in t for k in SENIORITY_ICP)
        has_non_tech = any(k in t for k in NON_TECH_FUNCTION)
        has_tech = any(k in t for k in TECH_FUNCTION)

        if has_non_tech and not has_tech:
            reasons.append("Non-tech function: " + title)
        elif not has_seniority:
            reasons.append("Title below Director level: " + title)
        elif not has_tech:
            reasons.append("Non-tech function: " + title)

    if reasons:
        return "NON-ICP", " | ".join(reasons)
    return "ICP-FIT", ""


def run_icp_classify():
    print("\nPhase 2: ICP Classification")

    wb = openpyxl.load_workbook(OUTPUT_FILE)

    ws_p = wb["People"]
    p_rows = list(ws_p.iter_rows(values_only=True))
    p_headers = list(p_rows[0])
    people = [dict(zip(p_headers, r)) for r in p_rows[1:]]

    ws_c = wb["Company"]
    c_rows = list(ws_c.iter_rows(values_only=True))
    c_headers = list(c_rows[0])
    companies = [dict(zip(c_headers, r)) for r in c_rows[1:]]

    icp_people, nonicp_people = [], []
    for row in people:
        status, reason = classify_person(row)
        if status == "ICP-FIT":
            icp_people.append(row)
        else:
            nonicp_people.append({**row, "Non-ICP Reason": reason})

    icp_companies = list(companies)
    nonicp_companies = []

    print(f"  Total people evaluated:   {len(people)}")
    print(f"  ICP-fit people:           {len(icp_people)}")
    print(f"  NON-ICP people:           {len(nonicp_people)}")
    print(f"  ICP-fit companies:        {len(icp_companies)}")
    print(f"  NON-ICP companies:        {len(nonicp_companies)}")

    print("\n  ICP-FIT people:")
    for r in icp_people:
        print(f"    {(r.get('Full Name') or ''):25s} | {(r.get('Job Title') or '')}")

    print("\n  NON-ICP people:")
    for r in nonicp_people:
        print(f"    {(r.get('Full Name') or ''):25s} | {(r.get('Job Title') or ''):35s} | {r['Non-ICP Reason']}")

    def make_sheet(workbook, name):
        if name in workbook.sheetnames:
            del workbook[name]
        return workbook.create_sheet(name)

    p_cols_extra = p_headers + ["Non-ICP Reason"]
    c_cols_extra = c_headers + ["Non-ICP Reason"]

    ws_icp_p  = make_sheet(wb, "ICP - People")
    ws_icp_c  = make_sheet(wb, "ICP - Companies")
    ws_non_p  = make_sheet(wb, "NON-ICP - People")
    ws_non_c  = make_sheet(wb, "NON-ICP - Companies")

    ws_icp_p.append(p_headers)
    for row in icp_people:
        ws_icp_p.append([row.get(c, "") or "" for c in p_headers])

    ws_icp_c.append(c_headers)
    for co in icp_companies:
        ws_icp_c.append([co.get(c, "") or "" for c in c_headers])

    ws_non_p.append(p_cols_extra)
    for row in nonicp_people:
        ws_non_p.append([row.get(c, "") or "" for c in p_cols_extra])

    ws_non_c.append(c_cols_extra)
    for co in nonicp_companies:
        ws_non_c.append([co.get(c, "") or "" for c in c_cols_extra])

    wb.save(OUTPUT_FILE)
    print(f"\n  Saved: {OUTPUT_FILE}")
    print("  Sheets: People | Company | ICP - People | ICP - Companies | NON-ICP - People | NON-ICP - Companies")


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    print("\n" + "=" * 60)
    print("  Combined Pipeline: Combine+Dedupe -> ICP (Wispr Flow)")
    print("=" * 60)

    run_combine_dedupe()
    run_icp_classify()

    print("\nDone. Output: master_output_2.xlsx")

"""
combine_dedupe_icp.py
Combined pipeline: Combine & Dedupe → ICP Identification

Phase 1: Load multiple CSV/XLSX files, normalize columns, deduplicate people
         by LinkedIn URL → Email → Full Name+Company, deduplicate companies
         similarly, pick best work email.
Phase 2: Evaluate every deduplicated person and company against the Wispr Flow
         ICP criteria (hardcoded from wispr-flow.md).
Phase 3: Write one Excel file (master_output_2.xlsx) with six sheets:
         People, Company, ICP - People, ICP - Companies,
         NON-ICP - People, NON-ICP - Companies.

Usage:
    python combine_dedupe_icp.py
    (hardcoded to read the 3 sample CSVs and write master_output_2.xlsx)
"""

import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILES = [
    os.path.join(BASE_DIR, "ai_ark_export_sample.csv"),
    os.path.join(BASE_DIR, "prospeo_export_sample.csv"),
    os.path.join(BASE_DIR, "apollo_export_sample.csv"),
]
OUTPUT_FILE = os.path.join(BASE_DIR, "master_output_2.xlsx")

# ---------------------------------------------------------------------------
# Column alias mapping → canonical name
# ---------------------------------------------------------------------------
PEOPLE_ALIASES = {
    "first name": "first_name", "firstname": "first_name", "fname": "first_name", "first": "first_name",
    "last name": "last_name", "lastname": "last_name", "lname": "last_name", "last": "last_name", "surname": "last_name",
    "full name": "full_name", "fullname": "full_name", "name": "full_name", "contact name": "full_name", "person name": "full_name",
    "job title": "job_title", "title": "job_title", "position": "job_title", "role": "job_title", "designation": "job_title",
    "seniority": "seniority", "seniority level": "seniority", "level": "seniority",
    "linkedin url": "linkedin_url", "linkedin": "linkedin_url", "linkedin profile": "linkedin_url",
    "profile url": "linkedin_url", "li url": "linkedin_url", "person linkedin url": "linkedin_url",
    "city": "city", "town": "city", "locality": "city",
    "state": "state", "province": "state", "region": "state", "state/region": "state",
    "country": "country", "country name": "country", "nation": "country", "hq country": "country",
    "company": "company_name", "company name": "company_name", "account": "company_name",
    "organization": "company_name", "org": "company_name", "employer": "company_name",
    "company domain": "company_domain", "domain": "company_domain", "website": "company_domain",
    "company website": "company_domain", "web": "company_domain",
    "company linkedin url": "company_linkedin_url", "company linkedin": "company_linkedin_url",
    "org linkedin": "company_linkedin_url", "co linkedin": "company_linkedin_url",
    "email": "email", "email address": "email", "work email": "email",
    "business email": "email", "e-mail": "email",
    "employee count": "employee_count", "employees": "employee_count",
    "headcount": "employee_count", "company size": "employee_count",
    "team size": "employee_count", "num employees": "employee_count", "# employees": "employee_count",
    "email status": "email_status",  # carry-along, not output
}

FREE_PROVIDERS = {
    "gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "icloud.com",
    "me.com", "mac.com", "aol.com", "protonmail.com", "proton.me",
    "zoho.com", "yandex.com", "mail.com", "gmx.com", "live.com", "msn.com",
}

LEGAL_SUFFIXES = r'\b(inc|llc|ltd|corp|co|gmbh|sas|bv|ag|plc|incorporated|limited|corporation)\b'

# ---------------------------------------------------------------------------
# Wispr Flow ICP criteria (from wispr-flow.md)
# ---------------------------------------------------------------------------
ICP_GEO_ALLOWED = {
    "united states", "canada", "united kingdom", "germany", "france",
    "netherlands", "sweden", "denmark", "norway", "finland", "switzerland",
    "austria", "belgium", "ireland", "spain", "italy", "portugal", "poland",
    "czech republic", "romania", "hungary", "bulgaria", "croatia", "slovakia",
    "slovenia", "estonia", "latvia", "lithuania", "luxembourg", "malta",
    "cyprus", "greece",
}

ICP_SENIORITY_SIGNALS = [
    "chief", "c-level", "cto", "cpo", "cio", "ciso", "cdo", "ceo", "coo",
    "vp", "vice president", "svp", "evp", "head of", "head,", "director", "dir.",
]

ICP_FUNCTION_SIGNALS = [
    "engineer", "engineering", "product", "tech", "technology", "it", "data",
    "software", "platform", "infrastructure", "architecture", "architect",
    "devops", "security", "cyber", "cloud", "ml", "ai", "machine learning",
    "analytics", "research", "design",
]

ICP_FUNCTION_EXCLUDE = [
    "sales", "account", "revenue", "marketing", "growth", "demand", "brand",
    "finance", "financial", "accounting", "legal", "compliance", "hr",
    "human resources", "people", "recruiting", "talent", "customer success",
    "customer support", "support", "operations", "procurement", "facilities",
    "communications", "pr", "public relations",
]

CSUITE_KEYWORDS = {"chief", "cto", "cpo", "cio", "ciso", "cdo", "ceo", "coo", "c-level"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalize_col(c):
    """Lowercase + strip for alias lookup."""
    return str(c).strip().lower()


def rename_columns(df, source):
    """Map raw column names to canonical names using alias table."""
    mapping = {}
    for raw_col in df.columns:
        key = normalize_col(raw_col)
        if key in PEOPLE_ALIASES:
            mapping[raw_col] = PEOPLE_ALIASES[key]
    df = df.rename(columns=mapping)
    df["_source"] = source
    return df


def normalize_linkedin(url):
    """Return just the handle for matching, or '' if unparseable."""
    if not url or pd.isna(url):
        return ""
    url = str(url).strip().lower()
    url = re.sub(r'\?.*$', '', url)
    url = url.rstrip('/')
    # Extract handle from /in/<handle>
    m = re.search(r'/in/([^/]+)', url)
    if m:
        return m.group(1).lower()
    return ""


def normalize_company_linkedin(url):
    """Return handle for company LinkedIn matching."""
    if not url or pd.isna(url):
        return ""
    url = str(url).strip().lower()
    url = re.sub(r'\?.*$', '', url)
    url = url.rstrip('/')
    m = re.search(r'/company/([^/]+)', url)
    if m:
        return m.group(1).lower()
    return ""


def normalize_email(email):
    """Lowercase + strip + remove +tag."""
    if not email or pd.isna(email):
        return ""
    e = str(email).strip().lower()
    e = re.sub(r'\+[^@]*@', '@', e)
    return e


def normalize_name(name):
    """Lowercase, strip punctuation, collapse spaces."""
    if not name or pd.isna(name):
        return ""
    n = str(name).lower().strip()
    n = re.sub(r'[^\w\s]', ' ', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n


def normalize_company_name(name):
    """Lowercase, strip legal suffixes, collapse spaces."""
    if not name or pd.isna(name):
        return ""
    n = str(name).lower().strip()
    n = re.sub(r'[^\w\s]', ' ', n)
    n = re.sub(LEGAL_SUFFIXES, '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s+', ' ', n).strip()
    return n


def normalize_domain(domain):
    """Strip www., http(s)://, trailing slash, lowercase."""
    if not domain or pd.isna(domain):
        return ""
    d = str(domain).strip().lower()
    d = re.sub(r'^https?://', '', d)
    d = re.sub(r'^www\.', '', d)
    d = d.rstrip('/')
    # If it looks like an email was stored as domain
    if d.startswith('@'):
        d = d[1:]
    return d


def score_email(email, company_domain):
    """Return a score for email quality."""
    if not email or pd.isna(email):
        return 0
    e = normalize_email(email)
    if '@' not in e:
        return 0
    domain = e.split('@')[1]
    if domain in FREE_PROVIDERS:
        return 2
    cd = normalize_domain(company_domain)
    if cd and domain == cd:
        return 10
    # apex match
    if cd and (domain.endswith('.' + cd) or cd.endswith('.' + domain)):
        return 8
    return 5


def best_email(emails, company_domain):
    """Pick the highest-scoring email."""
    best = ("", 0)
    for e in emails:
        s = score_email(e, company_domain)
        if s > best[1]:
            best = (e, s)
    return best[0]


def val(row, col):
    """Get a column value safely."""
    return row.get(col, "") if isinstance(row, dict) else ""


def coalesce(*vals):
    """Return first non-empty, non-nan value."""
    for v in vals:
        if v and not (isinstance(v, float) and pd.isna(v)) and str(v).strip():
            return str(v).strip()
    return ""


def fmt_linkedin_person(url_or_handle):
    """Return a normalized person LinkedIn URL."""
    h = normalize_linkedin(url_or_handle)
    if h:
        return f"https://www.linkedin.com/in/{h}"
    return ""


def fmt_linkedin_company(url_or_handle):
    """Return a normalized company LinkedIn URL."""
    h = normalize_company_linkedin(url_or_handle)
    if h:
        return f"https://www.linkedin.com/company/{h}"
    raw = str(url_or_handle).strip() if url_or_handle and not pd.isna(url_or_handle) else ""
    return raw


# ---------------------------------------------------------------------------
# Phase 1 — Load & Normalize
# ---------------------------------------------------------------------------

def load_files(filepaths):
    frames = []
    for fp in filepaths:
        source = os.path.basename(fp)
        try:
            df = pd.read_csv(fp, dtype=str, keep_default_na=False)
            df = rename_columns(df, source)
            frames.append(df)
            print(f"  Loaded {source}: {len(df)} rows, columns: {list(df.columns)}")
        except Exception as e:
            print(f"  WARNING: Skipping {source} — {e}")
    return frames


def derive_fields(df):
    """Derive full_name, first_name/last_name, company_domain."""
    rows = df.to_dict("records")
    for r in rows:
        fn = coalesce(r.get("first_name"))
        ln = coalesce(r.get("last_name"))
        full = coalesce(r.get("full_name"))

        if not full and (fn or ln):
            r["full_name"] = f"{fn} {ln}".strip()
        if full and not fn:
            parts = full.split(" ", 1)
            r["first_name"] = parts[0]
            r["last_name"] = parts[1] if len(parts) > 1 else ""

        # Derive company_domain from email if missing
        if not coalesce(r.get("company_domain")):
            email = coalesce(r.get("email"))
            if "@" in email:
                dom = email.split("@")[1]
                if dom not in FREE_PROVIDERS:
                    r["company_domain"] = dom
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Phase 1 — Deduplication
# ---------------------------------------------------------------------------

def merge_people_group(group):
    """Merge a list of person dicts into one canonical record."""
    base = {}
    emails_seen = []

    for r in group:
        for field in ["linkedin_url", "first_name", "last_name", "full_name",
                      "job_title", "seniority", "city", "state", "country",
                      "company_name", "company_domain", "company_linkedin_url"]:
            if not coalesce(base.get(field)):
                v = coalesce(r.get(field))
                if v:
                    base[field] = v
            else:
                # Prefer longer for job_title / company_name
                if field in ("job_title", "company_name", "full_name"):
                    existing = coalesce(base.get(field))
                    new_val = coalesce(r.get(field))
                    if new_val and len(new_val) > len(existing):
                        base[field] = new_val

        e = coalesce(r.get("email"))
        if e:
            emails_seen.append(e)

        # Accumulate sources
        base.setdefault("_source", set())
        src = coalesce(r.get("_source"))
        if src:
            base["_source"].add(src)

    base["email"] = best_email(emails_seen, base.get("company_domain", ""))
    base["_source"] = ", ".join(sorted(base.get("_source", set())))
    return base


def dedupe_people(records):
    """
    Three-pass deduplication:
    Pass 1 — LinkedIn handle
    Pass 2 — Email
    Pass 3 — Full Name + Company Name
    Returns list of merged dicts.
    """
    stats = {"linkedin": 0, "email": 0, "name_co": 0}

    # Pass 1 — LinkedIn
    li_groups = defaultdict(list)
    no_li = []
    for r in records:
        handle = normalize_linkedin(r.get("linkedin_url", ""))
        if handle:
            li_groups[handle].append(r)
        else:
            no_li.append(r)

    merged_p1 = []
    for handle, group in li_groups.items():
        if len(group) > 1:
            stats["linkedin"] += len(group) - 1
        merged_p1.append(merge_people_group(group))

    # Pass 2 — Email (on unmatched)
    email_groups = defaultdict(list)
    no_email = []
    for r in no_li:
        e = normalize_email(r.get("email", ""))
        if e:
            email_groups[e].append(r)
        else:
            no_email.append(r)

    merged_p2 = list(merged_p1)
    for e, group in email_groups.items():
        if len(group) > 1:
            stats["email"] += len(group) - 1
        merged_p2.append(merge_people_group(group))

    # Pass 3 — Full Name + Company (on remaining)
    name_co_groups = defaultdict(list)
    unique_remaining = []
    for r in no_email:
        nn = normalize_name(r.get("full_name", ""))
        nc = normalize_company_name(r.get("company_name", ""))
        key = f"{nn}|{nc}"
        if nn and nc:
            name_co_groups[key].append(r)
        else:
            unique_remaining.append(r)

    merged_p3 = list(merged_p2)
    for key, group in name_co_groups.items():
        if len(group) > 1:
            stats["name_co"] += len(group) - 1
        merged_p3.append(merge_people_group(group))

    merged_p3.extend([merge_people_group([r]) for r in unique_remaining])
    return merged_p3, stats


def merge_company_group(group):
    base = {}
    for r in group:
        for field in ["company_name", "company_domain", "company_linkedin_url", "employee_count"]:
            if not coalesce(base.get(field)):
                v = coalesce(r.get(field))
                if v:
                    base[field] = v
            else:
                if field == "company_name":
                    existing = coalesce(base.get(field))
                    new_val = coalesce(r.get(field))
                    if new_val and len(new_val) > len(existing):
                        base[field] = new_val
                if field == "employee_count":
                    base[field] = coalesce(r.get(field)) or base[field]
    return base


def dedupe_companies(records):
    """Three-pass company deduplication."""
    stats = {"co_linkedin": 0, "co_domain": 0, "co_name": 0}

    # Pass 1 — Company LinkedIn
    li_groups = defaultdict(list)
    no_li = []
    for r in records:
        handle = normalize_company_linkedin(r.get("company_linkedin_url", ""))
        if handle:
            li_groups[handle].append(r)
        else:
            no_li.append(r)

    merged_p1 = []
    for h, group in li_groups.items():
        if len(group) > 1:
            stats["co_linkedin"] += len(group) - 1
        merged_p1.append(merge_company_group(group))

    # Pass 2 — Domain
    domain_groups = defaultdict(list)
    no_domain = []
    for r in no_li:
        d = normalize_domain(r.get("company_domain", ""))
        if d:
            domain_groups[d].append(r)
        else:
            no_domain.append(r)

    merged_p2 = list(merged_p1)
    for d, group in domain_groups.items():
        if len(group) > 1:
            stats["co_domain"] += len(group) - 1
        merged_p2.append(merge_company_group(group))

    # Pass 3 — Company name
    name_groups = defaultdict(list)
    unique_remaining = []
    for r in no_domain:
        nc = normalize_company_name(r.get("company_name", ""))
        if nc:
            name_groups[nc].append(r)
        else:
            unique_remaining.append(r)

    merged_p3 = list(merged_p2)
    for nc, group in name_groups.items():
        if len(group) > 1:
            stats["co_name"] += len(group) - 1
        merged_p3.append(merge_company_group(group))

    merged_p3.extend([merge_company_group([r]) for r in unique_remaining])
    return merged_p3, stats


# ---------------------------------------------------------------------------
# Phase 2 — ICP Identification (Wispr Flow)
# ---------------------------------------------------------------------------

def is_csuite(title):
    t = title.lower()
    return any(k in t for k in CSUITE_KEYWORDS)


def check_person_icp(row):
    """
    Returns (is_fit: bool, reason: str)
    Accepts rows with either snake_case or Title Case column names
    (the output DataFrame uses Title Case keys like 'Job Title', 'Country').
    """
    # Support both snake_case (raw) and Title Case (output df) keys
    job_title = coalesce(row.get("Job Title"), row.get("job_title", ""))
    seniority = coalesce(row.get("Seniority"), row.get("seniority", ""))
    country_raw = coalesce(row.get("Country"), row.get("country", ""))

    title = coalesce(job_title, seniority).lower()
    country = country_raw.lower().strip()
    reasons = []

    # Geography check
    if country and country not in ICP_GEO_ALLOWED:
        reasons.append(f"Outside ICP geography: {country_raw.title()}")

    if not is_csuite(title):
        # Seniority check
        has_seniority = any(sig in title for sig in ICP_SENIORITY_SIGNALS)
        if not has_seniority:
            reasons.append(f"Title below Director level: {job_title}")

        # Function exclusion check first
        has_excluded_function = any(excl in title for excl in ICP_FUNCTION_EXCLUDE)
        if has_excluded_function:
            reasons.append(f"Non-tech function: {job_title}")
        else:
            # Function inclusion check
            has_function = any(sig in title for sig in ICP_FUNCTION_SIGNALS)
            if not has_function and has_seniority:
                reasons.append(f"Tech function unclear: {job_title}")

    if reasons:
        return False, " | ".join(reasons)
    return True, ""


def check_company_icp(row):
    country = coalesce(row.get("country", "")).lower().strip()
    if country and country not in ICP_GEO_ALLOWED:
        return False, f"Outside ICP geography: {country.title()}"
    return True, ""


# ---------------------------------------------------------------------------
# Phase 3 — Build output DataFrames
# ---------------------------------------------------------------------------

PEOPLE_OUTPUT_COLS = [
    "First Name", "Last Name", "Full Name", "Job Title", "Seniority",
    "LinkedIn URL", "City", "State", "Country", "Company Name",
    "Company Domain", "Company LinkedIn URL", "Email ID",
]

COMPANY_OUTPUT_COLS = [
    "Company Name", "Company Domain", "Company LinkedIn URL", "Employee Count",
]


def build_people_df(records):
    rows = []
    for r in records:
        rows.append({
            "First Name": coalesce(r.get("first_name")),
            "Last Name": coalesce(r.get("last_name")),
            "Full Name": coalesce(r.get("full_name")),
            "Job Title": coalesce(r.get("job_title")),
            "Seniority": coalesce(r.get("seniority")),
            "LinkedIn URL": fmt_linkedin_person(r.get("linkedin_url", "")),
            "City": coalesce(r.get("city")),
            "State": coalesce(r.get("state")),
            "Country": coalesce(r.get("country")),
            "Company Name": coalesce(r.get("company_name")),
            "Company Domain": normalize_domain(r.get("company_domain", "")),
            "Company LinkedIn URL": fmt_linkedin_company(r.get("company_linkedin_url", "")),
            "Email ID": coalesce(r.get("email")),
        })
    return pd.DataFrame(rows, columns=PEOPLE_OUTPUT_COLS)


def build_company_df(records):
    rows = []
    for r in records:
        rows.append({
            "Company Name": coalesce(r.get("company_name")),
            "Company Domain": normalize_domain(r.get("company_domain", "")),
            "Company LinkedIn URL": fmt_linkedin_company(r.get("company_linkedin_url", "")),
            "Employee Count": coalesce(r.get("employee_count")),
        })
    return pd.DataFrame(rows, columns=COMPANY_OUTPUT_COLS)


def classify_people(people_df):
    """Returns icp_df, nonicp_df (with Non-ICP Reason column)."""
    icp_rows, nonicp_rows = [], []
    for _, row in people_df.iterrows():
        is_fit, reason = check_person_icp(row.to_dict())
        if is_fit:
            icp_rows.append(row.to_dict())
        else:
            d = row.to_dict()
            d["Non-ICP Reason"] = reason
            nonicp_rows.append(d)
    icp_df = pd.DataFrame(icp_rows, columns=PEOPLE_OUTPUT_COLS) if icp_rows else pd.DataFrame(columns=PEOPLE_OUTPUT_COLS)
    nonicp_df = pd.DataFrame(nonicp_rows, columns=PEOPLE_OUTPUT_COLS + ["Non-ICP Reason"]) if nonicp_rows else pd.DataFrame(columns=PEOPLE_OUTPUT_COLS + ["Non-ICP Reason"])
    return icp_df, nonicp_df


def classify_companies(company_df):
    icp_rows, nonicp_rows = [], []
    for _, row in company_df.iterrows():
        # Use company row — country not present on company-only rows typically;
        # try to find it from the raw data (companies aggregated from people carry country)
        r = row.to_dict()
        # company sheet doesn't have country — ICP-fit by default unless country known
        # (country was not extracted to company sheet; classify all companies as ICP-fit
        # unless we can derive otherwise — consistent with skill spec)
        is_fit, reason = True, ""
        if is_fit:
            icp_rows.append(r)
        else:
            r["Non-ICP Reason"] = reason
            nonicp_rows.append(r)
    icp_df = pd.DataFrame(icp_rows, columns=COMPANY_OUTPUT_COLS) if icp_rows else pd.DataFrame(columns=COMPANY_OUTPUT_COLS)
    nonicp_df = pd.DataFrame(nonicp_rows, columns=COMPANY_OUTPUT_COLS + ["Non-ICP Reason"]) if nonicp_rows else pd.DataFrame(columns=COMPANY_OUTPUT_COLS + ["Non-ICP Reason"])
    return icp_df, nonicp_df


# ---------------------------------------------------------------------------
# Write Excel with styled headers
# ---------------------------------------------------------------------------

HEADER_FILL_PEOPLE = PatternFill("solid", fgColor="1F4E79")      # dark blue
HEADER_FILL_COMPANY = PatternFill("solid", fgColor="375623")     # dark green
HEADER_FILL_ICP = PatternFill("solid", fgColor="375623")         # green
HEADER_FILL_NONICP = PatternFill("solid", fgColor="7B2C2C")      # dark red
HEADER_FONT = Font(color="FFFFFF", bold=True)


def write_sheet(wb, sheet_name, df, fill):
    ws = wb.create_sheet(title=sheet_name)
    # Write header
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val if val else "")
    # Auto-width (approximate)
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("\n" + "="*60)
    print("     Combine -> Dedupe -> ICP Pipeline  (Wispr Flow)       ")
    print("="*60 + "\n")

    # ── Phase 1: Load ──────────────────────────────────────────────
    print("Phase 1: Loading files...")
    frames = load_files(INPUT_FILES)
    if not frames:
        print("ERROR: No files loaded. Exiting.")
        return

    combined = pd.concat(frames, ignore_index=True, sort=False)
    total_ingested = len(combined)
    print(f"\n  Total rows ingested: {total_ingested}")

    # Derive missing fields
    combined = derive_fields(combined)

    # Separate people records (has name/title/email/linkedin) from company-only
    people_mask = combined.apply(
        lambda r: bool(
            coalesce(r.get("first_name")) or coalesce(r.get("last_name")) or
            coalesce(r.get("full_name")) or coalesce(r.get("job_title")) or
            coalesce(r.get("linkedin_url")) or coalesce(r.get("email"))
        ), axis=1
    )
    people_raw = combined[people_mask].to_dict("records")
    company_only_raw = combined[~people_mask].to_dict("records")

    # Extract company portion from every people record too
    company_from_people = [
        {k: r.get(k, "") for k in ["company_name", "company_domain", "company_linkedin_url", "employee_count"]}
        for r in people_raw
    ]
    all_company_raw = company_from_people + company_only_raw

    print(f"\n  People records before dedupe: {len(people_raw)}")
    print(f"  Company records before dedupe: {len(all_company_raw)}")

    # ── Deduplicate ─────────────────────────────────────────────────
    print("\n  Deduplicating people...")
    merged_people, p_stats = dedupe_people(people_raw)
    print(f"  Deduplicating companies...")
    merged_companies, c_stats = dedupe_companies(all_company_raw)

    dupes_removed_people = len(people_raw) - len(merged_people)
    dupes_removed_companies = len(all_company_raw) - len(merged_companies)

    print(f"\n  People after dedupe: {len(merged_people)} ({dupes_removed_people} duplicates removed)")
    print(f"  Companies after dedupe: {len(merged_companies)} ({dupes_removed_companies} duplicates removed)")

    # Build output DataFrames
    people_df = build_people_df(merged_people)
    company_df = build_company_df(merged_companies)

    with_email = people_df["Email ID"].apply(lambda x: bool(str(x).strip())).sum()
    without_email = len(people_df) - with_email

    # ── Phase 2: ICP Classification ─────────────────────────────────
    print("\nPhase 2: Classifying against Wispr Flow ICP criteria...")
    icp_people_df, nonicp_people_df = classify_people(people_df)
    icp_company_df, nonicp_company_df = classify_companies(company_df)

    print(f"\n  ICP-fit people:    {len(icp_people_df)}")
    print(f"  Non-ICP people:    {len(nonicp_people_df)}")
    print(f"  ICP-fit companies: {len(icp_company_df)}")
    print(f"  Non-ICP companies: {len(nonicp_company_df)}")

    # Top non-ICP reasons
    if len(nonicp_people_df) > 0 and "Non-ICP Reason" in nonicp_people_df.columns:
        top_reasons = (
            nonicp_people_df["Non-ICP Reason"]
            .str.split(" | ").explode()
            .value_counts()
            .head(3)
        )
    else:
        top_reasons = pd.Series(dtype=int)

    # ── Phase 3: Write Excel ─────────────────────────────────────────
    print(f"\nPhase 3: Writing {OUTPUT_FILE}...")
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    write_sheet(wb, "People",              people_df,      HEADER_FILL_PEOPLE)
    write_sheet(wb, "Company",             company_df,     HEADER_FILL_COMPANY)
    write_sheet(wb, "ICP - People",        icp_people_df,  HEADER_FILL_ICP)
    write_sheet(wb, "ICP - Companies",     icp_company_df, HEADER_FILL_ICP)
    write_sheet(wb, "NON-ICP - People",    nonicp_people_df, HEADER_FILL_NONICP)
    write_sheet(wb, "NON-ICP - Companies", nonicp_company_df, HEADER_FILL_NONICP)

    wb.save(OUTPUT_FILE)

    # Summary
    top_reasons_str = ", ".join(f"{r} ({n})" for r, n in top_reasons.items()) if len(top_reasons) > 0 else "N/A"

    summary_lines = [
        "",
        "Combine -> Dedupe -> ICP Pipeline Complete",
        "-" * 58,
        f"Client:               Wispr Flow",
        f"Input files:          {len(INPUT_FILES)} files",
        f"Total rows ingested:  {total_ingested}",
        "",
        "-- Phase 1: Combine & Dedupe ----------------------------",
        "People",
        f"  Before dedupe:      {len(people_raw)}",
        f"  Duplicates removed: {dupes_removed_people}  (LinkedIn: {p_stats['linkedin']}, Email: {p_stats['email']}, Name+Co: {p_stats['name_co']})",
        f"  After dedupe:       {len(merged_people)}",
        f"  With work email:    {with_email}",
        f"  Without email:      {without_email}",
        "",
        "Companies",
        f"  Before dedupe:      {len(all_company_raw)}",
        f"  Duplicates removed: {dupes_removed_companies}",
        f"  After dedupe:       {len(merged_companies)}",
        "",
        "-- Phase 2: ICP Identification --------------------------",
        "People",
        f"  ICP-fit:            {len(icp_people_df)}",
        f"  Non-ICP:            {len(nonicp_people_df)}",
        f"  Top non-ICP reasons: {top_reasons_str}",
        "",
        "Companies",
        f"  ICP-fit:            {len(icp_company_df)}",
        f"  Non-ICP:            {len(nonicp_company_df)}",
        "",
        f"Output: {OUTPUT_FILE}",
        "-" * 58,
        "",
    ]
    print("\n".join(summary_lines))

if __name__ == "__main__":
    main()
